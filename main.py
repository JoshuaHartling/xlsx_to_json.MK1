from openpyxl import load_workbook
from jinja2 import Template

# active_sheet = "Santa_Barbara-US_CA"
active_sheet = "sheet1"
file_path = "C:/path/to/your/file"

# open workbook and worksheet
workbook = load_workbook(file_path, data_only=True)
worksheet = workbook[active_sheet]

# initialize data
data = []
adom = ''

for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
    # Get the keys from the first row of the worksheet
    keys = row
for row in worksheet.iter_rows(min_row=2, values_only=True):
    # turn None object into ""
    parameters = list(row)
    for index in range(len(parameters)):
        if parameters[index] is None:
            parameters[index] = ""
    # Create a dictionary for each row
    row_dict = dict(zip(keys, parameters))
    # Add the row dictionary to the main dictionary
    data.append(row_dict)

# assign adom
adom = data[0]["adom"]

# print statements
# print(adom)
# print(data)

json_template = """
{
    "adom": "{{adom}}",
    "variables": [
    {%- for row in rows %}
    {%- if row['variables_mapping_device'] is none and row['variables_mapping_vdom'] is none and row['variables_mapping_value'] is none %}
        {
            "name": "{{row['variables_name']}}",
            "value": "{{row['variables_value']}}"
        }{%- if not loop.last %},{% endif %}
    {%- else %}
        {
            "name": "{{row['variables_name']}}",
            "description": "{{row['variables_description']}}",
            "mapping": [
                {
                    "device": "{{row['variables_mapping_device']}}",
                    "vdom": "{{row['variables_mapping_vdom']}}",
                    "value": "{{row['variables_mapping_value']}}"
                }
            ]
        }{%- if not loop.last %},{% endif %}
    {%- endif %}
    {%- endfor %}
    ]
}
"""

for dictionary in data:
    template = Template(json_template)
    json_str = template.render(adom=adom, rows=data)
    print(json_str.strip())

# Save output to file 'output.json'
with open(f'{active_sheet}_output.json', 'w') as f:
    f.write(json_str.strip())